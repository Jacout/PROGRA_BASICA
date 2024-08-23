from openpyxl import Workbook
#trabajamos en un libro, luego sobre paginas, filas y columnas
#importamos workbook de openpyxl

libro = Workbook() #se instancia un objeto de tipo woorkbook para crear el libro
pagina = libro.active #cada libro se crea con una pagina, esta pagina es la activa es una propiedad

pagina1 = libro.create_sheet("Pagina1") #podemos crear paginas nuevas
pagina1.title = "Mi primera pagina" #Editar los nombres de las paginas (sheet)

print(libro.sheetnames) #retorna una lista con los nombres de las paginas


libro.save("Data base.xlsx") #Guardamos el libro de excel con el nombre


libro = load.workbook("Data base.xlsx")
