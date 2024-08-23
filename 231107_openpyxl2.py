from openpyxl import Workbook
from openpyxl import load_workbook
#para acceder al libro ya creado

try:
  libro =load_workbook("Data base.xlsx")
except FileNotFoundError:
  print(FileNotFoundError)
  exit()

lista=libro.sheetnames
print("Lista las paginas:",end=" ") #mostrar los nombres de las paginas

for a in lista:
  print(a, end=",")


if len(libro.sheetnames)>1: #se verifica si se tiene mas de una hoja 
  libro.active=1 #si si se activa con la segunda posicion 1 
  hoja=libro.active
else:
  hoja=libro.active #se activa con la que se tiene si no hay mas de una hoja
print("\n")
celda=hoja["A1"]
hoja["A1"]="Mi compa"
print("Accedemos a la celda A1 con el valor:", celda.value)

libro.save("Data base2.xlsx")