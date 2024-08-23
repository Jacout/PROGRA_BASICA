#generar archivos con dos pestañas, con datos personales y datos escolares
# Puedes usar la función open() para abrir un archivo en modo de escritura

from openpyxl import Workbook
from openpyxl import load_workbook


libro = Workbook()
datosPer = libro.active
datosPer.title = "Datos Personales"


#Llenar hoja de datos personales
datosPer["A1"]="Nombre"
datosPer["B1"]="Apellido"
datosPer["C1"]="Edad"
datosPer["D1"]="Dirección"
datosPer["E1"]="Ciudad"

#Relleno de datos
datosPer["A2"] = "Juan"
datosPer["B2"] = "Perez"
datosPer["C2"] = "25"
datosPer["D2"] = "Abriles 455"
datosPer["E2"] = "Guayaquil"
datosPer["A3"] = "Misael"
datosPer["B3"] = "Morales"
datosPer["C3"] = "20"
datosPer["D3"] = "Abril 16"
datosPer["E3"] = "Memerrey"

#Crear variable con objeto ya asignado de la nueva pestaña de datos escolares
datosEsc = libro.create_sheet("Datos Escolares")

#crear encabezados
datosEsc["A1"]="materia"
datosEsc["B1"]="nota"
datosEsc["C1"]="semestre"
datosEsc["D1"]="carrera"
datosEsc["E1"]="plan"

#Rellenar
datosEsc["A2"] = "Matematicas"
datosEsc["B2"] = "10"
datosEsc["C2"] = "2"
datosEsc["D2"] = "LMAD furro"
datosEsc["E2"] = "520"
datosEsc["A3"] = "Programacion basica"
datosEsc["B3"] = "9"
datosEsc["C3"] = "2"
datosEsc["D3"] = "LSTI"
datosEsc["E3"] = "430"
libro.save("Data.xlsx") #guarda y aqui8 acaba

import pandas as pd 

# Load the excel file
df = pd.read_excel("Data.xlsx")

# Escribir como csv
df.to_csv('Data.csv', index=False)