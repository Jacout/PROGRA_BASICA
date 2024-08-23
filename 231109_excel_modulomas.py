import random
from openpyxl import Workbook
from openpyxl import load_workbook

#cargamos un libro que ya ah sido creado
print("""Se abre un libro\n Si existe una excepcion, por ejemplo, si el libro no existe salimos del programa""")

try:
  libroseis = load_workbook("Data base.xlsx")
except FileNotFoundError:
  libroseis=Workbook()
input()
#enlistamos las paginas del libro y si hay mas paginas
#Sheet es la pagina 0
#Mi primera pagina 

print("Los nombres de las hojas son:", libroseis.sheetnames)
input()
if len(libroseis.sheetnames)>1:
  libroseis.active = 1
  hoja = libroseis.active
else:
  hoja =libroseis.active
print("La pagina es: ", hoja.title)
input()

#Accedemos a una celda en especifico
#De nuevo, se regresa un objeto a la variable celda
#Para manipular el objeto necesitamos metodoso o funciones
#El metoto value regresa el valor que contiene la celda

celda = hoja["A1"]
hoja["A1"] = "Hoy me levante para triunfar"
print("Accedemos a la cela A1 con valor: ", celda.value)
input()
#podemos acceder directamente a las celdas
hoja["A2"] = "Ya mañana solo Dios dira"
#(fila, columna, contenido)
hoja.cell(1,3,"YAAAAAAAAAAA nomas me miran pasar")
print("Los encabezados de las celdas son:")
print(hoja["A1"].value,end="\t")
print(hoja["B1"].value,end="\t")
print(hoja["C1"].value)

#Rango de celdas
#rango es de tipo tupla
rangoA = hoja["A2":"A6"]
rangoB = hoja["B2":"C6"]

print("Imprimimos los objetos de las celdas A2 a A6")
for celda in rangoA:
  for objeto in celda:
    print(celda, "contiene", objeto.value)
input()

print("Imprimimos los objetos de las celdas B2 a C6")
for fila in rangoB:
  for columna in fila:
    print(columna, "contiene", columna.value)
    
input()
#Llenado de celdas
c = 1
for fila in range(2,50):
  hoja.cell(fila,1,"192.168.1"+ str(c))
  c += 1

for fila in range(2,7):
  hoja.cell(fila,1,"Equipo "+chr(c))

for fila in range(2,7):
  lista = [str(random.randint(22,50)),
          str(random.randint(22,50)),
          str(random.randint(22,50))]
  hoja.cell(fila,3,str(lista))

#Impresion en pantalla de celdas
print("Impresion de todas las celdas")
for fila in hoja.values:
  for valor in fila:
    print(valor,end="\t")
  print("\n")

libroseis.save("Data base.xlsx") #siempre al final


#generar archivos con dos pestañas, con datos personales y datos escolares